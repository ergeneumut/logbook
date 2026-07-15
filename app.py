import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import io
import os
import base64
import datetime
import streamlit.components.v1 as components
from collections import defaultdict

# Sayfa Yapılandırması
st.set_page_config(page_title="Logbook Otomasyonu", layout="wide")

# Görselleri Base64 formatına çeviren yardımcı fonksiyon
def get_base64_image(file_path):
    if os.path.exists(file_path):
        with open(file_path, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode()
        ext = os.path.splitext(file_path)[1].lower()
        mime = "image/png" if ext == ".png" else "image/jpeg"
        return f"data:{mime};base64,{encoded_string}"
    return ""

# Görselleri klasörden oku
plane_base64 = get_base64_image("plane.jpg")
hangar_base64 = get_base64_image("background.jpg")

# CSS Temalandırma
bg_css = ""
if hangar_base64:
    bg_css = f"""
    .stApp {{
        background-image: url("{hangar_base64}");
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        background-attachment: fixed;
    }}
    [data-testid="stAppViewContainer"] {{
        background-color: rgba(15, 23, 42, 0.2) !important;
        backdrop-filter: none !important;
        -webkit-backdrop-filter: none !important;
    }}
    """

st.markdown(f"""
    <style>
    {bg_css}
    
    [data-testid="stHeader"] {{ background: transparent !important; }}
    
    .block-container {{
        background-color: rgba(255, 255, 255, 0.96) !important;
        padding: 3rem !important;
        border-radius: 16px;
        box-shadow: 0 15px 35px rgba(0, 0, 0, 0.4);
        max-width: 1100px !important;
        margin-top: 4vh !important;
        margin-bottom: 4vh !important;
    }}
    
    h1, h2, h3, h4, h5, h6, p, span, label, .stMarkdown, [data-testid="stWidgetLabel"] p {{
        color: #1e293b !important;
        text-shadow: none !important;
    }}
    
    [data-testid="stSidebar"] {{
        background-color: rgba(255, 255, 255, 0.98) !important;
        border-right: 1px solid rgba(0,0,0,0.1);
    }}
    [data-testid="stSidebar"] * {{
        color: #0f172a !important;
        text-shadow: none !important;
    }}
    
    [data-testid="stSidebar"] button {{
        background-color: #f1f5f9 !important;
        color: #0f172a !important;
        border: 1px solid #cbd5e1 !important;
    }}
    [data-testid="stSidebar"] button:hover {{
        background-color: #e2e8f0 !important;
        border-color: #94a3b8 !important;
    }}
    
    div[data-testid="column"] {{
        background-color: #ffffff !important;
        padding: 2.5rem !important;
        border-radius: 12px !important;
        border: 1px solid #e2e8f0 !important;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05) !important;
    }}
    
    div[data-testid="stRadio"] {{
        border-left: 6px solid transparent;
        padding: 15px !important;
        border-radius: 8px;
        transition: all 0.2s ease-in-out;
        background-color: rgba(255, 255, 255, 0.5);
        margin-bottom: 15px;
    }}
    
    .stRadio [data-testid="stMarkdownContainer"] {{
        font-size: 16px !important;
        font-weight: 500;
        color: #1e293b !important;
    }}
    
    [data-testid="stUploadDropzone"] {{
        background-color: #f8fafc !important;
        border: 2px dashed #3b82f6 !important;
    }}
    [data-testid="stUploadDropzone"] span, [data-testid="stUploadDropzone"] p {{
        text-shadow: none !important;
        color: #475569 !important;
    }}
    
    .stAlert {{
        background-color: #f8fafc !important;
        border: 1px solid #e2e8f0 !important;
    }}
    .stAlert p, .stAlert span, .stAlert div {{
        color: #1e293b !important;
        text-shadow: none !important;
    }}
    
    .stProgress > div > div > div > div {{
        background-color: #3b82f6 !important;
    }}
    
    div[data-testid="stDataFrame"] * {{
        text-shadow: none !important;
    }}
    </style>
""", unsafe_allow_html=True)

# State Yönetimi
if 'step' not in st.session_state:
    st.session_state.step = "upload"
if 'selected_jobs' not in st.session_state:
    st.session_state.selected_jobs = {}
if 'raw_data' not in st.session_state:
    st.session_state.raw_data = None
if 'unique_dates' not in st.session_state:
    st.session_state.unique_dates = []
if 'unique_months' not in st.session_state:
    st.session_state.unique_months = []
if 'current_month_idx' not in st.session_state:
    st.session_state.current_month_idx = 0
if 'original_file_bytes' not in st.session_state:
    st.session_state.original_file_bytes = None
if 'original_filename' not in st.session_state:
    st.session_state.original_filename = "Logbook_Raporu.xlsx"

def safe_get(lst, idx):
    return lst[idx] if idx < len(lst) else None

# Excel Verisini Yükleme Fonksiyonu
def load_excel_data(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    sheet = wb.active
    
    data = []
    last_valid_date = ""  # Akıllı Tarih Hafızası
    
    # 1. ve 2. satırlar başlık, veri 3'ten başlar
    for r_idx in range(3, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r_idx, column=c_idx).value for c_idx in range(1, sheet.max_column + 1)]
        
        # Tamamen boş satırları atla
        if not any(row_vals):
            continue
            
        date_val = safe_get(row_vals, 1)
        standard_date_str = ""
        
        if date_val:
            if hasattr(date_val, 'strftime'): 
                standard_date_str = date_val.strftime('%d.%m.%Y')
            else: 
                date_str = str(date_val).strip()
                if date_str:
                    parsed_dt = None
                    for fmt in ('%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                        try:
                            parsed_dt = pd.to_datetime(date_str, format=fmt)
                            break
                        except:
                            continue
                    if parsed_dt is not None and not pd.isna(parsed_dt):
                        standard_date_str = parsed_dt.strftime('%d.%m.%Y')
                    else:
                        standard_date_str = date_str
        
        # EĞER TARİH HÜCRESİ BOŞ BIRAKILMIŞSA (Aynı günün işleri için), BİR ÖNCEKİ TARİHİ OTOMATİK KULLAN
        if standard_date_str:
            last_valid_date = standard_date_str
        else:
            standard_date_str = last_valid_date
            
        # --- KUSURSUZ YENİ SÜTUN EŞLEŞTİRMELERİ ---
        privilege_val = safe_get(row_vals, 5)      # F Sütunu: 5. Privilege Used (Indis 5)
        ata_val = safe_get(row_vals, 20)            # U Sütunu: 10. ATA Chapter (Indis 20)
        check_type_val = safe_get(row_vals, 21)     # V Sütunu: Check Type (Indis 21)
        description_val = safe_get(row_vals, 22)    # W Sütunu: Description (Indis 22)
        duration_val = safe_get(row_vals, 23)       # X Sütunu: 12. Time Duration (Indis 23)
        ref_val = safe_get(row_vals, 24)            # Y Sütunu: 13. Maintenance Record Reference (Indis 24)
        wo_val = safe_get(row_vals, 25)             # Z Sütunu: W/O Number (Indis 25)

        # Dakika hesabı için duration_val kontrolü
        duration_mins = 0
        if duration_val:
            if isinstance(duration_val, (datetime.time, datetime.datetime)):
                duration_mins = duration_val.hour * 60 + duration_val.minute
            else:
                d_str = str(duration_val).strip()
                if ':' in d_str:
                    parts = d_str.split(':')
                    try: duration_mins = int(parts[0]) * 60 + int(parts[1])
                    except: pass
                else:
                    try: duration_mins = int(d_str)
                    except: pass

        data.append({
            "row_idx": r_idx,
            "no": safe_get(row_vals, 0),
            "date": standard_date_str,
            "location": safe_get(row_vals, 2),
            "fleet": safe_get(row_vals, 3),
            "privilege": str(privilege_val).strip() if privilege_val else "",
            "reg": safe_get(row_vals, 7),
            "ata": str(ata_val).strip() if ata_val else "",
            "check_type": str(check_type_val).strip() if check_type_val else "",
            "description": description_val,
            "duration": duration_val,
            "duration_mins": duration_mins,
            "ref": ref_val,
            "wo": wo_val
        })
    
    df = pd.DataFrame(data)
    df['date'] = df['date'].astype(str).str.strip()
    return df

def recalculate_dates_and_months(df):
    parsed_dates = []
    for d in df['date'].unique():
        if not d: continue
        try:
            dt = pd.to_datetime(d, format='%d.%m.%Y')
            parsed_dates.append((dt, d))
        except:
            parsed_dates.append((pd.Timestamp.max, d))
    parsed_dates.sort(key=lambda x: x[0])
    sorted_unique_dates = [x[1] for x in parsed_dates]

    unique_months = []
    for d in sorted_unique_dates:
        parts = d.split('.')
        if len(parts) == 3:
            yr, mn = parts[2], parts[1]
            if (yr, mn) not in unique_months:
                unique_months.append((yr, mn))
    return sorted_unique_dates, unique_months

def generate_output_excel(original_bytes, selected_row_indices, yellow_row_indices):
    wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    sheet = wb.active
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for r_idx in yellow_row_indices:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=r_idx, column=col).fill = yellow_fill

    max_row = sheet.max_row
    for r_idx in range(max_row, 2, -1):
        if r_idx not in selected_row_indices:
            sheet.delete_rows(r_idx)
            
    current_no = 1
    for r_idx in range(3, sheet.max_row + 1):
        sheet.cell(row=r_idx, column=1, value=current_no)
        current_no += 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ----------------- ADIM 1: DOSYA YÜKLEME -----------------
if st.session_state.step == "upload":
    st.title("✈️ Logbook Düzenleme Otomasyonu")
    st.subheader("Excel (.xlsx) Dosyanızı Yükleyin")
    
    uploaded_file = st.file_uploader("Dosya Seçin ve kısa bir süre bekleyin.", type=["xlsx"])
    
    if uploaded_file is not None:
        try:
            file_bytes = uploaded_file.getvalue()
            df = load_excel_data(file_bytes)
            
            st.session_state.original_file_bytes = file_bytes
            st.session_state.original_filename = uploaded_file.name
            st.session_state.raw_data = df
            
            st.session_state.step = "filter_data"
            st.rerun()
        except Exception as e:
            st.error(f"Hata: {e}. Lütfen doğru şablonda bir .xlsx dosyası yükleyin.")

# ----------------- ADIM 2: ÖN FİLTRELEME VE TEMİZLİK -----------------
elif st.session_state.step == "filter_data":
    st.title("🧹 Gereksiz İşleri Ayıklama (Filtreleme)")
    st.write("Listeden tamamen çıkarılması gereken, çok kısa süren veya belirli kriterlere (Yetki, ATA, Check Type) sahip olan işleri seçerek Excel tablosundan silebilirsiniz.")

    df = st.session_state.raw_data

    privileges = sorted([str(p) for p in df['privilege'].unique() if pd.notna(p) and str(p).strip()])
    atas = sorted([str(a) for a in df['ata'].unique() if pd.notna(a) and str(a).strip()])
    check_types = sorted([str(c) for c in df['check_type'].unique() if pd.notna(c) and str(c).strip()])

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### 🚫 Silinecek Kategoriler")
        sel_privs = st.multiselect("Silinecek Yetkilerini Seçin:", options=privileges)
        sel_atas = st.multiselect("Silinecek 'ATA Chapter' Bölümlerini Seçin:", options=atas)
        sel_checks = st.multiselect("Silinecek 'Check Type' Türlerini Seçin:", options=check_types)

    with col2:
        st.markdown("### ⏱️ Süre Filtresi")
        st.write("Süresi girilen dakikadan **daha az olan** tüm işler silinecektir.")
        min_minutes = st.number_input("Örn: 30:", min_value=0, max_value=600, value=0, step=15)

    mask = df['privilege'].isin(sel_privs) | df['ata'].isin(sel_atas) | df['check_type'].isin(sel_checks)
    if min_minutes > 0:
        mask = mask | (df['duration_mins'] < min_minutes)

    to_delete = df[mask]
    to_keep = df[~mask]

    if len(to_delete) > 0:
        st.error(f"⚠️ DİKKAT: Seçimlerinize göre toplam **{len(to_delete)} adet iş** Excel listesinden ve sistemden kalıcı olarak SİLİNECEKTİR.")
        st.markdown("**Silinecek İşlerin Ön İzlemesi:**")
        
        def color_red(val):
            return 'color: #ff4b4b; background-color: #fee2e2;'
            
        styled_df = to_delete[['date', 'wo', 'duration', 'privilege', 'ata', 'check_type', 'description']].style.map(color_red)
        st.dataframe(styled_df, use_container_width=True)

        st.warning("Yukarıdaki işlerin tablodan çıkarılmasını onaylıyor musunuz?")
        
        c1, c2 = st.columns(2)
        with c1:
            if st.button("🚨 Evet, Bu İşleri Sil ve Onayla", type="primary", use_container_width=True):
                if len(to_keep) == 0:
                    st.error("Tüm işleri sildiniz! Lütfen filtrelerinizi hafifletin.")
                else:
                    st.session_state.raw_data = to_keep
                    dates, months = recalculate_dates_and_months(to_keep)
                    st.session_state.unique_dates = dates
                    st.session_state.unique_months = months
                    
                    st.session_state.filter_msg = f"✅ Temizlik Başarılı! Başlangıçta {len(df)} iş/gün vardı, {len(to_delete)} tanesi silindi. Kalan iş/gün sayısı: {len(to_keep)}."
                    st.session_state.step = "choose_mode"
                    st.rerun()
        with c2:
            if st.button("❌ İptal Et ve Silmeden İlerle", use_container_width=True):
                dates, months = recalculate_dates_and_months(df)
                st.session_state.unique_dates = dates
                st.session_state.unique_months = months
                st.session_state.step = "choose_mode"
                st.rerun()
    else:
        st.info("ℹ️ Şu anki filtrelere göre silinecek hiçbir iş bulunmuyor. Farklı ayarlar deneyebilir veya doğrudan iş seçimine geçebilirsiniz.")
        st.write("---")
        if st.button("Temizliği Atla ve İş Seçimine Geç ➡️", use_container_width=True, type="primary"):
            dates, months = recalculate_dates_and_months(df)
            st.session_state.unique_dates = dates
            st.session_state.unique_months = months
            st.session_state.step = "choose_mode"
            st.rerun()


# ----------------- ADIM 3: SEÇİM MODU -----------------
elif st.session_state.step == "choose_mode":
    st.title("⚙️ İşlem Modunu Seçin")
    
    if 'filter_msg' in st.session_state:
        st.success(st.session_state.filter_msg)
        del st.session_state['filter_msg']
        
    st.write("Dosyanız işlemlere hazır. İşlemleri nasıl yapmak istersiniz?")
    
    col1, col2 = st.columns(2)
    with col1:
        st.success("⚡ 1. Seçenek: Otomatik (Hızlı)")
        st.write("**Otomatik ilk satırı seçer diğerlerini siler.**")
        st.write("Aynı günde sadece 1 iş seçimi yapar. Kalan günlerdeki ilk işi otomatik tutar, geri kalanları siler ve sizi doğrudan Sarı Boyama ekranına yönlendirir.")
        if st.button("Otomatik Seçimle İlerle", use_container_width=True, type="primary"):
            st.session_state.selected_jobs = {} # Seçimleri Temizle
            for date in st.session_state.unique_dates:
                daily_jobs = st.session_state.raw_data[st.session_state.raw_data['date'] == date]
                first_row_idx = daily_jobs.iloc[0]['row_idx']
                st.session_state.selected_jobs[date] = first_row_idx
            
            st.session_state.step = "select_samples"
            st.rerun()
            
    with col2:
        st.info("🖐️ 2. Seçenek: Manuel (Kontrollü)")
        st.write("**Tek tek elle iş seçerek devam edilir.**")
        st.write("Tüm günler için o güne ait işler listelenir. Hangi işin tabloda kalacağına okuyarak bizzat siz karar verirsiniz.")
        if st.button("Manuel Seçime Başla", use_container_width=True):
            st.session_state.selected_jobs = {} # Eski hafızayı tamamen temizle!
            st.session_state.current_month_idx = 0
            st.session_state.step = "select_daily"
            st.rerun()

# ----------------- ADIM 4: GRUPLANDIRILMIŞ SERİ MANUEL SEÇİM -----------------
elif st.session_state.step == "select_daily":
    df = st.session_state.raw_data
    dates = st.session_state.unique_dates
    unique_months = st.session_state.unique_months
    
    current_month_idx = st.session_state.current_month_idx
    active_year, active_month = unique_months[current_month_idx]
    
    month_dates = [d for d in dates if d.endswith(f"{active_month}.{active_year}")]
    
    components.html("""
    <script>
    const doc = window.parent.document;
    window.parent._activeGroupIdx = 0;
    
    function highlightActiveGroup() {
        const groups = Array.from(doc.querySelectorAll('div[data-testid="stRadio"]'));
        const activeIdx = window.parent._activeGroupIdx || 0;
        
        groups.forEach((group, idx) => {
            if (idx === activeIdx) {
                group.style.borderLeftColor = "#3b82f6";
                group.style.backgroundColor = "#f1f5f9";
                group.scrollIntoView({ behavior: 'smooth', block: 'center' });
                
                const firstInput = group.querySelector('input[type="radio"]');
                if (firstInput) firstInput.focus();
            } else {
                group.style.borderLeftColor = "transparent";
                group.style.backgroundColor = "rgba(255, 255, 255, 0.5)";
            }
        });
    }
    
    function setupClickListeners() {
        const groups = Array.from(doc.querySelectorAll('div[data-testid="stRadio"]'));
        groups.forEach((group, idx) => {
            if (!group._hasClickListener) {
                group._hasClickListener = true;
                group.addEventListener('click', function() {
                    window.parent._activeGroupIdx = idx;
                    highlightActiveGroup();
                });
            }
        });
    }

    const initTimer = setInterval(function() {
        const groups = doc.querySelectorAll('div[data-testid="stRadio"]');
        if (groups.length) {
            highlightActiveGroup();
            setupClickListeners();
            clearInterval(initTimer);
        }
    }, 100);

    if (!window.parent._groupedKeyboardListenerAdded) {
        window.parent._groupedKeyboardListenerAdded = true;
        
        window.parent.document.addEventListener('keydown', function(e) {
            if (e.target.tagName === 'INPUT' && e.target.type === 'text') return;
            
            const groups = Array.from(window.parent.document.querySelectorAll('div[data-testid="stRadio"]'));
            if (groups.length === 0) return;
            
            let idx = window.parent._activeGroupIdx || 0;
            const activeGroup = groups[idx];
            if (!activeGroup) return;
            
            const labels = Array.from(activeGroup.querySelectorAll('label'));
            const inputs = Array.from(activeGroup.querySelectorAll('input[type="radio"]'));
            const checkedIndex = inputs.findIndex(r => r.checked);
            
            if (e.key === 'ArrowDown') {
                e.preventDefault();
                if (checkedIndex < labels.length - 1) {
                    labels[checkedIndex + 1].click();
                }
            } 
            else if (e.key === 'ArrowUp') {
                e.preventDefault();
                if (checkedIndex > 0) {
                    labels[checkedIndex - 1].click();
                }
            }
            else if (e.key === 'Enter') {
                e.preventDefault(); 
                
                if (idx < groups.length - 1) {
                    idx++;
                    window.parent._activeGroupIdx = idx;
                    
                    groups.forEach((g, gIdx) => {
                        if (gIdx === idx) {
                            g.style.borderLeftColor = "#3b82f6";
                            g.style.backgroundColor = "#f1f5f9";
                            g.scrollIntoView({ behavior: 'smooth', block: 'center' });
                            
                            const firstInput = g.querySelector('input[type="radio"]');
                            if (firstInput) firstInput.focus();
                        } else {
                            g.style.borderLeftColor = "transparent";
                            g.style.backgroundColor = "rgba(255, 255, 255, 0.5)";
                        }
                    });
                } else {
                    let submitBtn = window.parent.document.querySelector('button[data-testid="baseButton-secondaryFormSubmit"]');
                    if (!submitBtn) {
                        const buttons = Array.from(window.parent.document.querySelectorAll('button'));
                        submitBtn = buttons.find(btn => btn.textContent.includes('Kaydet ve İlerle') || btn.textContent.includes('Örnek İş Seçimine Geç'));
                    }
                    if (submitBtn) {
                        submitBtn.focus();
                        submitBtn.click();
                    }
                }
            }
        }, true); 
    }
    </script>
    """, height=0)

    st.sidebar.title("📅 Aylık Seçici")
    st.sidebar.write("Geçmek istediğiniz aya doğrudan tıklayabilirsiniz:")
    
    MONTH_NAMES = {
        1: 'Ocak', 2: 'Şubat', 3: 'Mart', 4: 'Nisan', 5: 'Mayıs', 6: 'Haziran',
        7: 'Temmuz', 8: 'Ağustos', 9: 'Eylül', 10: 'Ekim', 11: 'Kasım', 12: 'Aralık'
    }

    sidebar_groups = defaultdict(list)
    for idx, (yr, mn) in enumerate(unique_months):
        sidebar_groups[yr].append((idx, mn))
        
    for yr in sorted(sidebar_groups.keys()):
        is_year_expanded = (yr == active_year)
        with st.sidebar.expander(f"📁 {yr} Yılı", expanded=is_year_expanded):
            for idx, mn in sidebar_groups[yr]:
                month_name = MONTH_NAMES.get(int(mn), mn)
                month_dates_list = [d for d in dates if d.endswith(f"{mn}.{yr}")]
                is_month_complete = all(d in st.session_state.selected_jobs for d in month_dates_list)
                status_icon = "🟢" if is_month_complete else "⚪"
                
                btn_label = f"{status_icon} {month_name}"
                if idx == current_month_idx:
                    btn_label = f"👉 {month_name} (Aktif)"
                    if st.button(btn_label, key=f"side_btn_{yr}_{mn}", use_container_width=True, type="primary"):
                        st.session_state.current_month_idx = idx
                        st.rerun()
                else:
                    if st.button(btn_label, key=f"side_btn_{yr}_{mn}", use_container_width=True):
                        st.session_state.current_month_idx = idx
                        st.rerun()

    st.title("⚡ Seri İş Seçim Ekranı")
    progress = len(st.session_state.selected_jobs) / len(dates)
    st.progress(progress)
    st.subheader(f"Dönem: {MONTH_NAMES.get(int(active_month), active_month)} {active_year}")
    
    st.info("⌨️ **İş seçimleri arasında gezinmek için **Aşağı/Yukarı Ok** tuşlarını kullanın. Seçimi yapıp bir alt güne kaymak için **Enter** tuşuna basın.")
    
    with st.form(key=f"month_form_{active_year}_{active_month}"):
        form_selections = {}
        
        for d_str in month_dates:
            daily_jobs = df[df['date'] == d_str]
            options = {}
            for _, row in daily_jobs.iterrows():
                # Benzersiz etiket yapısı eklendi
                label = f"W/O: {row['wo']} | Ref: {row['ref']} | Süre: {row['duration']} | Tanım: {row['description']}"
                options[row['row_idx']] = label
                
            default_val = list(options.keys())[0]
            if d_str in st.session_state.selected_jobs:
                default_val = st.session_state.selected_jobs[d_str]
                
            selected_row = st.radio(
                f"📌 Gün: {d_str}",
                options=list(options.keys()),
                format_func=lambda x: options[x],
                index=list(options.keys()).index(default_val),
                key=f"radio_{d_str}"
            )
            form_selections[d_str] = selected_row
            st.write("---")
            
        is_last_month = (current_month_idx == len(unique_months) - 1)
        submit_label = "🎉 Örnek İş Seçimine Geç" if is_last_month else "Bu Ayın Seçimlerini Kaydet ve İlerle ➡️ (Enter)"
        
        submitted = st.form_submit_button(submit_label, use_container_width=True, type="primary")
        
        if submitted:
            for d_str, r_idx in form_selections.items():
                st.session_state.selected_jobs[d_str] = r_idx
                
            if is_last_month:
                st.session_state.step = "select_samples"
            else:
                st.session_state.current_month_idx += 1
            st.rerun()

    if current_month_idx > 0:
        if st.button("⬅️ Önceki Aya Geri Dön", use_container_width=True):
            st.session_state.current_month_idx -= 1
            st.rerun()

# ----------------- ADIM 5: ÖRNEK İŞ SEÇİMİ (SARI BOYAMA) -----------------
elif st.session_state.step == "select_samples":
    st.title("🎯 Örnek İşlerin Belirlenmesi")
    
    selected_indices = list(st.session_state.selected_jobs.values())
    df = st.session_state.raw_data
    selected_df = df[df['row_idx'].isin(selected_indices)].copy()
    
    selected_df['temp_sort_date'] = pd.to_datetime(selected_df['date'], format='%d.%m.%Y', errors='coerce')
    selected_df = selected_df.sort_values('temp_sort_date')
    
    st.write(f"Toplam **{len(selected_df)}** gün/iş filtrelendi.")
    st.info("Örnek olarak gösterilecek (sarıya boyanacak) işleri sol taraftaki kutucuklardan seçin.")
    
    selected_df['Sarı Boya (Örnek İş)'] = False
    
    edited_df = st.data_editor(
        selected_df[['Sarı Boya (Örnek İş)', 'date', 'reg', 'wo', 'ref', 'duration', 'description', 'fleet', 'location', 'row_idx']],
        disabled=['date', 'location', 'fleet', 'reg', 'description', 'duration', 'wo', 'row_idx'],
        column_config={
            "Sarı Boya (Örnek İş)": st.column_config.CheckboxColumn("Sarı Boya?", default=False)
        },
        hide_index=True,
        use_container_width=True
    )
    
    yellow_rows = edited_df[edited_df['Sarı Boya (Örnek İş)'] == True]['row_idx'].tolist()
    st.metric("Seçilen Örnek İş Sayısı (Sarı Boyanacaklar)", f"{len(yellow_rows)}")
    
    col_b1, col_b2 = st.columns([1, 1])
    with col_b1:
        if st.button("⬅️ Geri Dön ve Düzenleme Modu Seç", use_container_width=True):
            st.session_state.step = "choose_mode"
            st.rerun()
            
    with col_b2:
        if st.button("Excel Dosyasını Hazırla 📥", type="primary", use_container_width=True):
            with st.spinner("Exceli hazırlıyorum, azıcık sabret :)..."):
                excel_data = generate_output_excel(
                    st.session_state.original_file_bytes, 
                    selected_indices, 
                    yellow_rows
                )
                st.session_state.final_excel = excel_data
                st.session_state.step = "download"
                st.rerun()

# ----------------- ADIM 6: İNDİRME EKRANI -----------------
elif st.session_state.step == "download":
    st.title("🏆 Raporunuz Hazır!")
    
    if plane_base64:
        plane_html = f'<img class="thy-plane-img" src="{plane_base64}">'
        plane_css = """
        .thy-plane-img {
            width: 380px;
            height: auto;
            filter: drop-shadow(-10px 15px 15px rgba(0,0,0,0.5));
        }
        """
    else:
        plane_html = '<span style="font-size: 130px; filter: drop-shadow(5px 10px 10px rgba(0,0,0,0.3)); display: inline-block; transform: scaleX(-1);">✈️</span><span style="font-size: 50px; font-weight: 900; font-style: italic; color: #D61C22; background: white; padding: 5px 20px; border-radius: 12px; border: 4px solid #D61C22; margin-left: 15px; box-shadow: -3px 5px 15px rgba(0,0,0,0.4);">THY</span>'
        plane_css = ""

    components.html(f"""
    <script>
    const parentDoc = window.parent.document;
    if (!parentDoc.getElementById('thy-anim')) {{
        const style = parentDoc.createElement('style');
        style.innerHTML = `
        @keyframes flyPlane {{
            0% {{ left: 120vw; top: 75vh; transform: rotate(15deg) scale(0.6); opacity: 0; }}
            10% {{ opacity: 1; }}
            90% {{ opacity: 1; }}
            100% {{ left: -40vw; top: 10vh; transform: rotate(5deg) scale(1.3); opacity: 0; }}
        }}
        .thy-plane-wrapper {{
            position: fixed;
            z-index: 9999999;
            pointer-events: none;
            animation: flyPlane 5.5s cubic-bezier(0.25, 1, 0.5, 1) forwards;
            display: flex;
            align-items: center;
        }}
        {plane_css}
        `;
        parentDoc.head.appendChild(style);
        
        const plane = parentDoc.createElement('div');
        plane.id = 'thy-anim';
        plane.className = 'thy-plane-wrapper';
        plane.innerHTML = '{plane_html}';
        parentDoc.body.appendChild(plane);
        
        setTimeout(() => {{
            if (plane.parentNode) plane.parentNode.removeChild(plane);
        }}, 6000);
    }}
    </script>
    """, height=0)

    st.success("Tebrikler! Seçtiğin ayarlara ve temizlik filtrelerine göre dosya başarıyla oluşturuldu. Tabloyu ve içeriklerin doğruluğunu mutlaka kontrol ediniz.Şimdi kahve molası :)")
    
    st.download_button(
        label="📥 Düzenlenmiş Excel Dosyasını İndir",
        data=st.session_state.final_excel,
        file_name=st.session_state.original_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    
    if st.button("🔄 Yeni Bir Dosya İle Yeniden Başla", use_container_width=True):
        st.session_state.clear()
        st.rerun()
